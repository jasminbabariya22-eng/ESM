from sqlalchemy.orm import Session, joinedload
from datetime import datetime, timezone
from sqlalchemy.inspection import inspect
from sqlalchemy import func
import pandas as pd
import io

import math
from openpyxl.styles import PatternFill, Alignment,Font

from app.models.department import Department
from app.models.risk_register import RiskRegister
from app.models.risk_description import RiskDescription
from app.models.risk_treatment import RiskTreatment
from app.models.risk_action_followup import RiskActionFollowup

from app.models.risk_register_hist import RiskRegisterHist
from app.models.risk_description_hist import RiskDescriptionHist
from app.models.risk_treatment_hist import RiskTreatmentHist

from app.models.mst_status import Status
from app.models.user import User


# Generate Risk ID
def generate_risk_id(db: Session, dept_id: int):

    dept = db.query(Department).filter(
        Department.id == dept_id
    ).with_for_update().first()

    if not dept:
        raise Exception("Department not found")

    dept.last_risk_number += 1
    number = dept.last_risk_number

    risk_id = f"{dept.dept_short_name}-{str(number).zfill(4)}"

    return risk_id


# Type Conversion
def to_int(val):
    return int(val) if val not in [None, ""] else None


def to_float(val):
    return float(val) if val not in [None, ""] else None


def to_datetime(val):
    from datetime import datetime
    return datetime.fromisoformat(val) if val not in [None, ""] else None


def model_to_dict(obj):
    return {c.key: getattr(obj, c.key) for c in inspect(obj).mapper.column_attrs}


# Get Status ID
def get_status_id(db: Session, status_name: str):
    status = db.query(Status).filter(
        Status.status_name == status_name,
        Status.is_deleted == 0
    ).first()

    if not status:
        raise Exception(f"Status '{status_name}' not found in mst_status")

    return status.id

# ---------
# CREATE OR UPDATE RISK
# ---------

def create_update_risk(db: Session, data, current_user):

    try:

        if not data or not data.risk_register:
            raise ValueError("risk_register is required")

        register_data = data.risk_register
        desc_data = data.risk_description
        treatments = data.risk_treatments or []

        # CREATE OR UPDATE RISK REGISTER

        if to_int(register_data.risk_register_id) == 0:

            risk_id = generate_risk_id(db, to_int(register_data.dept_id))

            risk = RiskRegister(
                risk_id=risk_id,
                risk_name=register_data.risk_name,
                dept_id=to_int(register_data.dept_id),
                risk_owner_id=to_int(register_data.risk_owner_id),
                risk_co_owner_id=to_int(register_data.risk_co_owner_id),
                financial_year=register_data.financial_year,
                risk_status=to_int(register_data.risk_status),
                risk_progress=to_float(register_data.risk_progress),
                created_by=current_user["id"],
                created_on=datetime.now(timezone.utc),
                is_active=0,
                is_deleted=0
            )

            db.add(risk)
            db.flush()

        else:

            risk = db.query(RiskRegister).filter(
                RiskRegister.risk_register_id == to_int(register_data.risk_register_id)
            ).first()

            if not risk:
                raise ValueError("RiskRegister not found")

            risk.risk_name = register_data.risk_name
            risk.dept_id = to_int(register_data.dept_id)
            risk.risk_owner_id = to_int(register_data.risk_owner_id)
            risk.risk_co_owner_id = to_int(register_data.risk_co_owner_id)
            risk.financial_year = register_data.financial_year
            risk.risk_status = to_int(register_data.risk_status)
            risk.risk_progress = to_float(register_data.risk_progress)

            risk.modified_by = current_user["id"]
            risk.modified_on = datetime.now(timezone.utc)


        # HISTORY - RISK REGISTER

        hist_register = RiskRegisterHist(
            risk_register_id=risk.risk_register_id,
            risk_id=risk.risk_id,
            risk_name=risk.risk_name,
            dept_id=risk.dept_id,
            risk_owner_id=risk.risk_owner_id,
            risk_co_owner_id=risk.risk_co_owner_id,
            financial_year=risk.financial_year,
            risk_status=risk.risk_status,
            risk_progress=risk.risk_progress,
            created_by=risk.created_by,
            created_on=risk.created_on,
            modified_by=risk.modified_by,
            modified_on=risk.modified_on,
            is_active=risk.is_active,
            is_deleted=risk.is_deleted
        )

        db.add(hist_register)


        # RISK DESCRIPTION

        description = None

        if desc_data and any([
            desc_data.risk_description not in [None, ""],
            desc_data.mitigation not in [None, ""],
            to_int(desc_data.inherent_risk_likelihood_id) not in [None, 0],
            to_int(desc_data.inherent_risk_impact_id) not in [None, 0],
            to_int(desc_data.current_risk_likelihood_id) not in [None, 0],
            to_int(desc_data.current_risk_impact_id) not in [None, 0]
        ]):

            if to_int(desc_data.risk_description_id) == 0:

                description = RiskDescription(
                    risk_register_id=risk.risk_register_id,
                    risk_id=risk.risk_id,
                    risk_description=desc_data.risk_description,
                    inherent_risk_likelihood_id=to_int(desc_data.inherent_risk_likelihood_id),
                    inherent_risk_impact_id=to_int(desc_data.inherent_risk_impact_id),
                    mitigation=desc_data.mitigation,
                    current_risk_likelihood_id=to_int(desc_data.current_risk_likelihood_id),
                    current_risk_impact_id=to_int(desc_data.current_risk_impact_id),
                    created_by=current_user["id"],
                    created_on=datetime.now(timezone.utc),
                    is_deleted=0
                )

                db.add(description)
                db.flush()

            # Update Risk Description
            else:

                description = db.query(RiskDescription).filter(
                    RiskDescription.risk_description_id == to_int(desc_data.risk_description_id)
                ).first()

                if not description:
                    raise ValueError("RiskDescription not found")

                description.risk_description = desc_data.risk_description
                description.inherent_risk_likelihood_id = to_int(desc_data.inherent_risk_likelihood_id)
                description.inherent_risk_impact_id = to_int(desc_data.inherent_risk_impact_id)
                description.mitigation = desc_data.mitigation
                description.current_risk_likelihood_id = to_int(desc_data.current_risk_likelihood_id)
                description.current_risk_impact_id = to_int(desc_data.current_risk_impact_id)

                description.modified_by = current_user["id"]
                description.modified_on = datetime.now(timezone.utc)


            # HISTORY DESCRIPTION

            hist_desc = RiskDescriptionHist(
                risk_description_id=description.risk_description_id,
                risk_register_id=description.risk_register_id,
                risk_id=description.risk_id,
                risk_description=description.risk_description,
                inherent_risk_likelihood_id=description.inherent_risk_likelihood_id,
                inherent_risk_impact_id=description.inherent_risk_impact_id,
                mitigation=description.mitigation,
                current_risk_likelihood_id=description.current_risk_likelihood_id,
                current_risk_impact_id=description.current_risk_impact_id,
                created_by=description.created_by,
                created_on=description.created_on,
                modified_by=description.modified_by,
                modified_on=description.modified_on,
                is_deleted=description.is_deleted
            )

            db.add(hist_desc)

        
        # RISK TREATMENTS

        saved_treatments = []

        if description is not None:

            if desc_data and to_int(desc_data.risk_description_id) > 0:
                db.query(RiskTreatment).filter(
                    RiskTreatment.risk_description_id == description.risk_description_id
                ).delete()

            for treatment in treatments:

                new_treatment = RiskTreatment(
                    risk_register_id=risk.risk_register_id,
                    risk_description_id=description.risk_description_id,
                    risk_id=risk.risk_id,
                    action_plan=treatment.action_plan,
                    action_owner_id=to_int(treatment.action_owner_id),
                    target_date=to_datetime(treatment.target_date),
                    progress=to_float(treatment.progress),
                    action_status_id=to_int(treatment.action_status_id),
                    next_followup_date=to_datetime(treatment.next_followup_date),
                    created_by=current_user["id"],
                    created_on=datetime.now(timezone.utc),
                    is_deleted=0
                )

                db.add(new_treatment)
                db.flush()

                saved_treatments.append(new_treatment)
                
                
                # AUTO UPDATE RISK STATUS → IN PROGRESS
                if len(saved_treatments) > 0:
                    in_progress_status = get_status_id(db, "In Progress")
                    risk.risk_status = in_progress_status

                hist_treatment = RiskTreatmentHist(
                    risk_treatment_id=new_treatment.risk_treatment_id,
                    risk_description_id=new_treatment.risk_description_id,
                    risk_register_id=new_treatment.risk_register_id,
                    risk_id=new_treatment.risk_id,
                    action_plan=new_treatment.action_plan,
                    action_owner_id=new_treatment.action_owner_id,
                    target_date=new_treatment.target_date,
                    progress=new_treatment.progress,
                    action_status_id=new_treatment.action_status_id,
                    next_followup_date=new_treatment.next_followup_date,
                    created_by=new_treatment.created_by,
                    created_on=new_treatment.created_on,
                    is_deleted=new_treatment.is_deleted
                )

                db.add(hist_treatment)

        
        # AUTO CALCULATE RISK PROGRESS
        
        avg_progress = db.query(
            func.avg(RiskTreatment.progress)
        ).filter(
            RiskTreatment.risk_register_id == risk.risk_register_id,
            RiskTreatment.is_deleted == 0
        ).scalar()

        risk.risk_progress = round(avg_progress or 0, 2)

        db.commit()

        return {
            "risk_register": model_to_dict(risk),
            "risk_description": model_to_dict(description) if description else None,
            "risk_treatments": [model_to_dict(t) for t in saved_treatments]
        }

    except Exception as e:
        db.rollback()
        raise e


# Risk get by User (Optinal API)
def get_risk_by_user(db, user_id):

    risks = db.query(RiskRegister).filter(
        RiskRegister.risk_owner_id == user_id,
        RiskRegister.is_deleted == 0
    ).all()

    result = []

    for risk in risks:

        description = db.query(RiskDescription).filter(
            RiskDescription.risk_register_id == risk.risk_register_id
        ).first()

        treatments = []

        if description:
            treatments = db.query(RiskTreatment).filter(
                RiskTreatment.risk_description_id == description.risk_description_id
            ).all()

        result.append({
            "risk_register_id": risk.risk_register_id,
            "risk_id": risk.risk_id,
            "risk_name": risk.risk_name,
            "financial_year": risk.financial_year,
            "risk_status": risk.risk_status,
            "risk_progress": risk.risk_progress,
            "description": description,
            "treatments": treatments
        })

    return result




def to_dict(obj, model=None,prefix=None):
    def format_key(key):
        return f"{prefix}{key}" if prefix else key
    
    # If object exists
    if obj is not None:
        return {
            format_key(c.key): getattr(obj, c.key)
            for c in inspect(obj).mapper.column_attrs
        }

    # If object is None → return all columns as None
    if model is not None:
        return {
            format_key(c.key): None
            for c in inspect(model).mapper.column_attrs
        }
    return {}

def get_color(score):
    
    if score <= 4:
        return "#4CAF50"
    elif score <= 9:
        return "#FFEB3B"
    elif score <= 16:
        return "#FF9800"
    else:
        return "#F44336"


#-----------
# Get Risk by id (Assign id)
#----------
def get_risk_by_id(db, id):
    impact_map = {1:"A",2:"B",3:"C",4:"D",5:"E"}
    try:
        query = db.query(
                RiskRegister,
                RiskDescription,
                RiskTreatment
            ).join(
                RiskDescription,
                RiskRegister.risk_register_id == RiskDescription.risk_register_id
            ).join(
                RiskTreatment,
                RiskDescription.risk_description_id == RiskTreatment.risk_description_id
            ).filter(
                RiskRegister.is_deleted == 0
            )
            
        if id:
            query = query.filter(RiskTreatment.action_owner_id == id)

        records = query.order_by(
                RiskRegister.risk_register_id,
                RiskDescription.risk_description_id,
                RiskTreatment.risk_treatment_id
            ).all()    
        

        #result = [ {**to_dict(rr), **to_dict(rd), **to_dict(rt)} for rr, rd, rt in records ]
        result = []
        for rr, rd, rt in records:
            risk_owner_name = None
            if rr.risk_owner:
                risk_owner_name = rr.risk_owner.log_id
                
            risk_co_owner_name = None
            if rr.risk_co_owner:
                risk_co_owner_name = rr.risk_co_owner.log_id
                
            risk_status_name = None
            if rt.status:
                risk_status_name = rt.status.status_name
            
            
            likelihood = rd.inherent_risk_likelihood_id
            impact = rd.inherent_risk_impact_id
            current_likelihood = rd.current_risk_likelihood_id
            current_impact = rd.current_risk_impact_id

            inherent_color_str = None
            inherent_color_code = None
            current_color_str = None
            current_color_code = None

            if likelihood and impact:
                inherent_color_str = get_color(likelihood*impact)
                inherent_color_code = f"{likelihood}{impact_map.get(impact)}"
                current_color_str = get_color(current_likelihood*current_impact)
                current_color_code = f"{current_likelihood}{impact_map.get(current_impact)}"

            result.append({
                **to_dict(rr),
                **to_dict(rd),
                **to_dict(rt),
                "inherent_color_str": inherent_color_str,
                "inherent_color_code" : inherent_color_code,
                
                "current_color_str": current_color_str,
                "current_color_code" : current_color_code,
                
                "risk_owner_name" : risk_owner_name,
                "risk_co_owner_name" : risk_co_owner_name,
                "risk_status_name": risk_status_name
            })

        return result
    except Exception as e:
        raise e
    
    
#-----------
# Risk LIST from Department id
#----------
def get_risk_by_dept(db, dept_id):  
    impact_map = {1:"A",2:"B",3:"C",4:"D",5:"E"}
    try:
        query = db.query(
                    RiskRegister,
                    RiskDescription
                ).outerjoin(
                    RiskDescription,
                    RiskRegister.risk_register_id == RiskDescription.risk_register_id
                )
                
        query = query.filter(RiskRegister.is_deleted == 0)

        if dept_id:
            query = query.filter( RiskRegister.dept_id == dept_id )


        records = query.order_by(
                RiskRegister.risk_register_id,
                RiskDescription.risk_description_id
            ).all()    
        

        result = []
        for rr, rd in records:
            risk_owner_name = None
            if rr.risk_owner:
                risk_owner_name = rr.risk_owner.log_id
            
            risk_status_name = None
            if rr.status:
                risk_status_name = rr.status.status_name

            likelihood = None
            impact = None
            current_likelihood = None
            current_impact = None
            inherent_color_str = None
            inherent_color_code = None
            current_color_str = None
            current_color_code = None
           
            if rd is not None:
                if rd.inherent_risk_likelihood_id:
                    likelihood = rd.inherent_risk_likelihood_id
                if rd.inherent_risk_impact_id:
                    impact = rd.inherent_risk_impact_id

                if rd.current_risk_likelihood_id:
                    current_likelihood = rd.current_risk_likelihood_id
                if rd.current_risk_impact_id:
                    current_impact = rd.current_risk_impact_id


                if likelihood and impact:
                    inherent_color_str = get_color(likelihood*impact)
                    inherent_color_code = f"{likelihood}{impact_map.get(impact)}"

                if current_likelihood and current_impact:
                    current_color_str = get_color(current_likelihood*current_impact)
                    current_color_code = f"{current_likelihood}{impact_map.get(current_impact)}"

            result.append({
                **to_dict(rr),
                **to_dict(rd, RiskDescription, prefix="rd_"),
                "inherent_color_str": inherent_color_str,
                "inherent_color_code" : inherent_color_code,
                "current_color_str": current_color_str,
                "current_color_code" : current_color_code,
                "risk_owner_name" : risk_owner_name,
                "risk_status_name" : risk_status_name
            })

        return result
    except Exception as e:
        raise e


#-----------
# Risk Register by risk_id
#------------

def get_risk_by_risk_id(db, risk_id):
    impact_map = {1:"A",2:"B",3:"C",4:"D",5:"E"}

    try:
        risks = (
            db.query(RiskRegister)
            .options(
                joinedload(RiskRegister.risk_owner),
                joinedload(RiskRegister.risk_co_owner),
                joinedload(RiskRegister.status),

                joinedload(RiskRegister.function_head_status),
                joinedload(RiskRegister.risk_head_status),
                joinedload(RiskRegister.risk_manager_status),

                joinedload(RiskRegister.risk_function_head_approval_by_name),
                joinedload(RiskRegister.risk_head_approval_by_name),
                joinedload(RiskRegister.risk_manager_approval_by_name),

                joinedload(RiskRegister.risk_descriptions)
                .joinedload(RiskDescription.treatments)
                .joinedload(RiskTreatment.action_owner),

                joinedload(RiskRegister.risk_descriptions)
                .joinedload(RiskDescription.treatments)
                .joinedload(RiskTreatment.status)
            )
            .filter(
                RiskRegister.risk_register_id == risk_id,
                RiskRegister.is_deleted == 0
            )
            .all()
        )

        result = []

        for rr in risks:

            risk_dict = to_dict(rr)

            # ---------- Owner & Status ----------
            risk_dict["rd_risk_owner_name"] = rr.risk_owner.log_id if rr.risk_owner else None
            risk_dict["rd_risk_co_owner_name"] = rr.risk_co_owner.log_id if rr.risk_co_owner else None
            risk_dict["rd_risk_status_name"] = rr.status.status_name if rr.status else None

            # ---------- Function Head ----------
            risk_dict["risk_function_head_approval_status_name"] = (
                rr.function_head_status.status_name if rr.function_head_status else None
            )
            risk_dict["risk_function_head_approval_by_name"] = (
                rr.risk_function_head_approval_by_name.log_id if rr.risk_function_head_approval_by_name else None
            )

            # ---------- Risk Head ----------
            risk_dict["risk_head_approval_status_name"] = (
                rr.risk_head_status.status_name if rr.risk_head_status else None
            )
            risk_dict["risk_head_approval_by_name"] = (
                rr.risk_head_approval_by_name.log_id if rr.risk_head_approval_by_name else None
            )

            # ---------- Risk Manager ----------
            risk_dict["risk_manager_approval_status_name"] = (
                rr.risk_manager_status.status_name if rr.risk_manager_status else None
            )
            risk_dict["risk_manager_approval_by_name"] = (
                rr.risk_manager_approval_by_name.log_id if rr.risk_manager_approval_by_name else None
            )

            # ---------- Risk Descriptions ----------
            risk_desc_list = []

            for rd in rr.risk_descriptions:

                likelihood = rd.inherent_risk_likelihood_id
                impact = rd.inherent_risk_impact_id
                current_likelihood = rd.current_risk_likelihood_id
                current_impact = rd.current_risk_impact_id

                inherent_color_str = None
                inherent_color_code = None
                current_color_str = None
                current_color_code = None

                if likelihood and impact:
                    inherent_color_str = get_color(likelihood * impact)
                    inherent_color_code = f"{likelihood}{impact_map.get(impact)}"

                if current_likelihood and current_impact:
                    current_color_str = get_color(current_likelihood * current_impact)
                    current_color_code = f"{current_likelihood}{impact_map.get(current_impact)}"

                # ---------- Treatments ----------
                treatments_list = []

                for rt in rd.treatments:
                    treatments_list.append({
                        **to_dict(rt),
                        "risk_owner_name": rt.action_owner.log_id if rt.action_owner else None,
                        "risk_co_owner_name": rt.action_owner.log_id if rt.action_owner else None,
                        "risk_status_name": rt.status.status_name if rt.status else None
                    })

                rd_dict = {
                    **to_dict(rd),
                    "inherent_color_str": inherent_color_str,
                    "inherent_color_code": inherent_color_code,
                    "current_color_str": current_color_str,
                    "current_color_code": current_color_code,
                    "treatments": treatments_list
                }

                risk_desc_list.append(rd_dict)

            risk_dict["risk_descriptions"] = risk_desc_list
            result.append(risk_dict)

        return result

    except Exception as e:
        raise e
    
    
# Risk Register by risk_description_id
# def get_risk_by_description_id(db, description_id):
#     try:

#         risk_description = (
#             db.query(RiskDescription)
#             .options(
                
#                 joinedload(RiskDescription.treatments)
#                 .joinedload(RiskTreatment.status)
#                 .load_only(Status.status_name),

#                 joinedload(RiskDescription.treatments)
#                 .joinedload(RiskTreatment.action_owner)
#                 .load_only(User.log_id)
#                 )
#             .filter(
#                 RiskDescription.risk_description_id == description_id,
#                 RiskDescription.is_deleted == 0
#             )
#             .first()
#         )

#         result = risk_description.__dict__.copy()

#         treatments_list = []

#         for t in risk_description.treatments:
#             treatment_dict = t.__dict__.copy()

#             # remove SQLAlchemy internal state
#             treatment_dict.pop("_sa_instance_state", None)

#             # add status_name
#             treatment_dict["risk_status_name"] = (
#                 t.status.status_name if t.status else None
#             )
#             treatment_dict["risk_owner_name"] = (
#                 t.action_owner.log_id if t.action_owner else None
#             )

#             # remove nested status object
#             treatment_dict.pop("status", None)
#             treatment_dict.pop("action_owner", None)

#             treatments_list.append(treatment_dict)

#         result["treatments"] = treatments_list
#         result.pop("_sa_instance_state", None)
#         return result

#     except Exception as e:
#         raise e


#----------
# Risk Register by risk_description_id
#----------

def get_risk_by_description_id(db, description_id):

    impact_map = {1:"A",2:"B",3:"C",4:"D",5:"E"}

    try:

        descriptions = (
            db.query(RiskDescription)
            .options(
                joinedload(RiskDescription.treatments)
                .joinedload(RiskTreatment.action_owner),

                joinedload(RiskDescription.treatments)
                .joinedload(RiskTreatment.status),

                joinedload(RiskDescription.risk_register)
            )
            .filter(
                RiskDescription.risk_description_id == description_id,
                RiskDescription.is_deleted == 0
            )
            .all()
        )

        result = []

        for rd in descriptions:

            likelihood = rd.inherent_risk_likelihood_id
            impact = rd.inherent_risk_impact_id
            current_likelihood = rd.current_risk_likelihood_id
            current_impact = rd.current_risk_impact_id

            inherent_color_str = None
            inherent_color_code = None
            current_color_str = None
            current_color_code = None

            if likelihood and impact:
                inherent_color_str = get_color(likelihood * impact)
                inherent_color_code = f"{likelihood}{impact_map.get(impact)}"

            if current_likelihood and current_impact:
                current_color_str = get_color(current_likelihood * current_impact)
                current_color_code = f"{current_likelihood}{impact_map.get(current_impact)}"

            # Treatments
            treatments_list = []

            for rt in rd.treatments:
                treatments_list.append({
                    **to_dict(rt),
                    "risk_owner_name": rt.action_owner.log_id if rt.action_owner else None,
                    "risk_co_owner_name": rt.action_owner.log_id if rt.action_owner else None,
                    "risk_status_name": rt.status.status_name if rt.status else None
                })

            rd_dict = {
                **to_dict(rd),
                "inherent_color_str": inherent_color_str,
                "inherent_color_code": inherent_color_code,
                "current_color_str": current_color_str,
                "current_color_code": current_color_code,
                "treatments": treatments_list
            }

            # Risk Register info
            risk_dict = {
                **to_dict(rd.risk_register),
                "risk_descriptions": [rd_dict]
            }

            result.append(risk_dict)

        return result

    except Exception as e:
        raise e
    
    
#---------    
# Download Risk data in Excel format
#----------

def calculate_row_height(text, column_width):
    if not text:
        return 15
    
    # Estimate characters per line
    chars_per_line = column_width * 1.2
    
    # Estimate number of lines
    lines = math.ceil(len(str(text)) / chars_per_line)
    
    # Excel default row height ≈ 15
    return max(15, lines * 15)

    
def get_risk_data_excel(db,dept_id):
    impact_map = {1:"A",2:"B",3:"C",4:"D",5:"E"}
    green_value = {"1A", "2A", "3A", "1B", "2B","1C"}
    yellow_value = {"4A", "5A", "3B", "3C", "2D","2C","1D","1E"}
    amber_value = {"2E", "3E", "3D", "4C", "4B","5B"}
    red_value = {"4E", "5E", "4D", "5D", "5C"}
    
    green_fill = PatternFill(start_color="4CAF50",
                            end_color="4CAF50",
                            fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB3B",
                            end_color="FFEB3B",
                            fill_type="solid")
    amber_fill = PatternFill(start_color="FF9800",
                            end_color="FF9800",
                            fill_type="solid")
    red_fill = PatternFill(start_color="F44336",
                            end_color="F44336",
                            fill_type="solid")
    try:
        query = db.query(RiskRegister).options(
                joinedload(RiskRegister.risk_descriptions)
                .joinedload(RiskDescription.treatments)
            )
        if dept_id:
            query = query.filter(RiskRegister.dept_id == dept_id)
        query = query.filter(RiskRegister.is_deleted == 0, RiskRegister.dept_id > 0).order_by(RiskRegister.risk_register_id)
        risks = query.all()

        department_rows  = {}
        merge_ranges = {}
        current_rows = {}

        #current_row = 2
        for risk in risks:
            first_risk = True
            risk_owner_name = ""
            dept_name = "UNKNOWN"
            if risk.risk_owner :
                risk_owner_name = risk.risk_owner.log_id

            if risk.department:
                dept_name = risk.department.dept_short_name

            if dept_name not in department_rows:
                department_rows[dept_name] = []
                merge_ranges[dept_name] = []
                current_rows[dept_name] = 2

            start_row = current_rows[dept_name]
            descriptions = risk.risk_descriptions if risk.risk_descriptions else [None]
            
            for desc in descriptions:

                first_desc = True
                likelihood = desc.inherent_risk_likelihood_id if desc else None
                impact = desc.inherent_risk_impact_id if desc else None
                current_likelihood = desc.current_risk_likelihood_id if desc else None
                current_impact = desc.current_risk_impact_id if desc else None

                inherent_color_str = None
                inherent_color_code = None
                current_color_str = None
                current_color_code = None

                if likelihood and impact:
                    inherent_color_str = get_color(likelihood*impact)
                    inherent_color_code = f"{likelihood}{impact_map.get(impact)}"
                if current_likelihood and current_impact:
                    current_color_str = get_color(current_likelihood*current_impact)
                    current_color_code = f"{current_likelihood}{impact_map.get(current_impact)}"

                treatments = desc.treatments if desc and desc.treatments else [None]
                for treatment in treatments:

                    department_rows[dept_name].append({
                        "Risk ID": risk.risk_id if first_risk else "",
                        "Risk Name" : risk.risk_name if first_risk else "",
                        "Risk Description": desc.risk_description if desc and first_desc else "",
                        "Inherent Risk Level" : inherent_color_code if first_desc else "",
                        "Current Mitigation": desc.mitigation if desc and first_desc else "",
                        "Current Risk Level" : current_color_code if first_desc else "",
                        "Risk Owner" : risk_owner_name if first_risk else "",
                        "Action Owner" : treatment.action_owner.log_id if treatment and treatment.action_owner else "",
                        "Risk Treatment": treatment.action_plan if treatment else "",
                        "Target Date" :treatment.target_date.date() if treatment and treatment.target_date else ""
                        
                    })

                    first_risk = False
                    first_desc = False
                    current_rows[dept_name] += 1
            end_row = current_rows[dept_name] - 1
            if end_row >= start_row:
                merge_ranges[dept_name].append((start_row, end_row))

        #df = pd.DataFrame(rows)

        output = io.BytesIO()
        wrap_alignment = Alignment(
            wrap_text=True,
            vertical="center"
        )

        merge_alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            for dept_name, rows in department_rows.items():
                df = pd.DataFrame(rows)
                sheet_name = dept_name[:31]
                df.to_excel(writer, index=False, sheet_name=sheet_name)

                workbook = writer.book
                worksheet = writer.sheets[sheet_name]
            

                # Apply color logic
                for row in range(2, len(df) + 2):   # Excel row index starts at 1
                    cell = worksheet[f"D{row}"]     # Column D = Inherent Risk Level

                    if cell.value in green_value:
                        cell.fill = green_fill
                    elif cell.value in amber_value:
                        cell.fill = amber_fill
                    elif cell.value in yellow_value:
                        cell.fill = yellow_fill
                    elif cell.value in red_value:
                        cell.fill = red_fill

                    cell_f = worksheet[f"F{row}"]     # Column D = Inherent Risk Level

                    if cell_f.value in green_value:
                        cell_f.fill = green_fill
                    elif cell_f.value in amber_value:
                        cell_f.fill = amber_fill
                    elif cell_f.value in yellow_value:
                        cell_f.fill = yellow_fill
                    elif cell_f.value in red_value:
                        cell_f.fill = red_fill

                # Column Widths
                column_widths = {
                    "A": 15,   # Risk ID (A)
                    "B" :20, # Risk Name (B)
                    "C": 50,   # Risk Description (C)
                    "D": 10,   # Inherent Risk Level (D)
                    "E": 70,    # Mitigation (E)
                    "F": 10, # Current risk level (F)
                    "G": 15, # Risk Owner (G)
                    "H": 15, # Action Owner (H)
                    "I": 70, # Treatment (I)
                    "J": 15  # Target Date (J)
                }

                for col, width in column_widths.items():
                    worksheet.column_dimensions[col].width = width

                
                for start, end in merge_ranges[dept_name]:
                    if start != end:

                        worksheet.merge_cells(f"A{start}:A{end}")  # Risk ID
                        worksheet.merge_cells(f"B{start}:B{end}")  # Risk Name
                        # worksheet.merge_cells(f"C{start}:C{end}")  # Risk Description
                        # worksheet.merge_cells(f"D{start}:D{end}")  # Inherent Risk Level
                        # worksheet.merge_cells(f"E{start}:E{end}")  # Current Mitigation
                        # worksheet.merge_cells(f"F{start}:F{end}")  # Current Risk Level
                        worksheet.merge_cells(f"G{start}:G{end}")  # Risk Owner
                    
                    worksheet[f"A{start}"].alignment = merge_alignment
                    worksheet[f"B{start}"].alignment = merge_alignment
                    worksheet[f"C{start}"].alignment = merge_alignment
                    worksheet[f"D{start}"].alignment = merge_alignment
                    worksheet[f"E{start}"].alignment = merge_alignment
                    worksheet[f"F{start}"].alignment = merge_alignment
                    worksheet[f"G{start}"].alignment = merge_alignment

                # Wrap Text for all cells
                for row in range(2, len(df)+2):
                    worksheet[f"C{row}"].alignment = wrap_alignment   # Description
                    worksheet[f"E{row}"].alignment = wrap_alignment   # Mitigation
                    worksheet[f"I{row}"].alignment = wrap_alignment   # Treatment
                    
                    desc = worksheet[f"C{row}"].value
                    mitigation = worksheet[f"E{row}"].value
                    treatment = worksheet[f"I{row}"].value

                    height_desc = calculate_row_height(desc, column_widths["C"])
                    height_mit = calculate_row_height(mitigation, column_widths["E"])
                    height_treat = calculate_row_height(treatment, column_widths["I"])

                    worksheet.row_dimensions[row].height = max(height_desc, height_mit, height_treat)

                # Wrap Header
                header_alignment = Alignment(
                    wrap_text=True,
                    horizontal="center",
                    vertical="center"
                )
                
                header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                header_font = Font(bold=True)

                for cell in worksheet[1]:
                    cell.alignment = header_alignment
                    cell.fill = header_fill
                    cell.font = header_font

        output.seek(0)

        return output
    except Exception as e:
        raise e


# def fetch_risk_data(db, dept_id):
#     query = db.query(RiskRegister).options(
#         joinedload(RiskRegister.risk_descriptions)
#         .joinedload(RiskDescription.treatments)
#     )

#     if dept_id:
#         query = query.filter(RiskRegister.dept_id == dept_id)

#     query = query.filter(
#         RiskRegister.is_deleted == 0,
#         RiskRegister.dept_id > 0
#     ).order_by(RiskRegister.risk_register_id)

#     return query.all()

# def calculate_risk_levels(desc, impact_map):
#     def get_code_and_color(likelihood, impact):
#         if not likelihood or not impact:
#             return None, None

#         code = f"{likelihood}{impact_map.get(impact)}"
#         color = get_color(likelihood * impact)

#         return code, color

#     inherent_code, inherent_color = get_code_and_color(
#         getattr(desc, "inherent_risk_likelihood_id", None),
#         getattr(desc, "inherent_risk_impact_id", None)
#     )

#     current_code, current_color = get_code_and_color(
#         getattr(desc, "current_risk_likelihood_id", None),
#         getattr(desc, "current_risk_impact_id", None)
#     )

#     return {
#         "inherent_code": inherent_code,
#         "inherent_color": inherent_color,
#         "current_code": current_code,
#         "current_color": current_color
#     }
    

# def format_excel(department_rows, merge_ranges):
#     output = io.BytesIO()

#     green_value = {"1A", "2A", "3A", "1B", "2B","1C"}
#     yellow_value = {"4A", "5A", "3B", "3C", "2D","2C","1D","1E"}
#     amber_value = {"2E", "3E", "3D", "4C", "4B","5B"}
#     red_value = {"4E", "5E", "4D", "5D", "5C"}

#     with pd.ExcelWriter(output, engine="openpyxl") as writer:

#         for dept_name, rows in department_rows.items():
#             df = pd.DataFrame(rows)
#             sheet_name = dept_name[:31]
#             df.to_excel(writer, index=False, sheet_name=sheet_name)

#             worksheet = writer.sheets[sheet_name]

#             # 🎨 Color logic
#             for row in range(2, len(df)+2):
#                 for col in ["D", "F"]:
#                     cell = worksheet[f"{col}{row}"]

#                     if cell.value in green_value:
#                         cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
#                     elif cell.value in amber_value:
#                         cell.fill = PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid")
#                     elif cell.value in yellow_value:
#                         cell.fill = PatternFill(start_color="FFEB3B", end_color="FFEB3B", fill_type="solid")
#                     elif cell.value in red_value:
#                         cell.fill = PatternFill(start_color="F44336", end_color="F44336", fill_type="solid")

#     output.seek(0)
#     return output

# def get_risk_data_excel(db, dept_id):
#     impact_map = {1:"A",2:"B",3:"C",4:"D",5:"E"}

#     risks = fetch_risk_data(db, dept_id)

#     department_rows = {}
#     merge_ranges = {}
#     current_rows = {}

#     for risk in risks:
#         dept_name = risk.department.dept_short_name if risk.department else "UNKNOWN"

#         if dept_name not in department_rows:
#             department_rows[dept_name] = []
#             merge_ranges[dept_name] = []
#             current_rows[dept_name] = 2

#         start_row = current_rows[dept_name]

#         descriptions = risk.risk_descriptions or [None]

#         for desc in descriptions:
#             risk_levels = calculate_risk_levels(desc, impact_map)

#             treatments = desc.treatments if desc and desc.treatments else [None]

#             for treatment in treatments:
#                 department_rows[dept_name].append({
#                     "Risk ID": risk.risk_id,
#                     "Risk Name": risk.risk_name,
#                     "Risk Description": getattr(desc, "risk_description", ""),
#                     "Inherent Risk Level": risk_levels["inherent_code"],
#                     "Current Risk Level": risk_levels["current_code"],
#                     "Risk Owner": risk.risk_owner.log_id if risk.risk_owner else ""
#                 })

#                 current_rows[dept_name] += 1

#         merge_ranges[dept_name].append((start_row, current_rows[dept_name] - 1))

#     return format_excel(department_rows, merge_ranges)
    

#---------
# Get Followups by reference_id
#---------

def get_followups_by_reference_id(db, reference_id):
    try:

        followups = (
            db.query(RiskActionFollowup)
            .options(
                joinedload(RiskActionFollowup.created_user)
                .load_only(User.log_id),
                
                joinedload(RiskActionFollowup.status_master)
                .load_only(Status.status_name)
            )
            .filter(
                RiskActionFollowup.reference_id == reference_id
            )
            .all()
        )

        result = []

        for followup in followups:

            followup_dict = followup.__dict__.copy()

            # remove SQLAlchemy internal state
            followup_dict.pop("_sa_instance_state", None)
            followup_dict.pop("file_data", None)

            # add user name
            followup_dict["risk_owner_name"] = (
                followup.created_user.log_id if followup.created_user else None
            )
            
            # add status name
            followup_dict["risk_status_name"] = (
                followup.status_master.status_name if followup.status_master else None
            )

            # remove relationship object
            followup_dict.pop("created_user", None)
            followup_dict.pop("status_master", None)

            result.append(followup_dict)

        return result

    except Exception as e:
        raise e